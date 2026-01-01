#!/usr/bin/env python3
"""
Time & Energy Audit Tracker
Inspired by Dan Martell's time tracking strategy
Pops up every 15 minutes to log activities with color coding and dollar values

Copyright (c) 2026 Arty McLabin
"""

import tkinter as tk
from tkinter import messagebox
import threading
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import os
import re
import sys

class TimeTracker:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Time Tracker - Dan Martell Method")
        self.root.geometry("700x300")

        # Get script directory for Excel file
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_file = os.path.join(self.script_dir, "time_audit.xlsx")
        self.setup_excel()

        # Timer setup - single timer, no duplication
        self.timer_minutes = 15
        self.timer_start_time = None
        self.timer_running = False
        self.has_popped = False

        # Color mapping for Excel
        self.color_fills = {
            'green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'red': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
            'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }

        self.setup_ui()
        self.reset_timer()
        self.start_timer_display()

    def setup_excel(self):
        """Initialize Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Time Audit"

            # Headers - separate Date and Time columns
            headers = ['Date', 'Time', 'Energy', 'Value', 'Activity']
            ws.append(headers)

            # Make headers bold
            for cell in ws[1]:
                cell.font = Font(bold=True)

            wb.save(self.excel_file)

    def setup_ui(self):
        """Create the UI"""
        # Instructions
        instructions = tk.Label(
            self.root,
            text='Format: color (g/r/w or green/red/white) + dollars + activity (any order, spaces optional)\nExamples: "g$$ activity" or "$$r activity" or "white activity$$" | Press Enter to save',
            font=('Consolas', 9),
            justify=tk.LEFT,
            bg='#1e1e1e',
            fg='#00ff00',
            padx=10,
            pady=10
        )
        instructions.pack(fill=tk.X)

        # Timer display
        self.timer_label = tk.Label(
            self.root,
            text="Next reminder in: 15:00",
            font=('Consolas', 12, 'bold'),
            bg='#1e1e1e',
            fg='#ffff00'
        )
        self.timer_label.pack(pady=5)

        # Status label (shows save confirmation)
        self.status_label = tk.Label(
            self.root,
            text="",
            font=('Consolas', 10),
            bg='#1e1e1e',
            fg='#00ff00',
            height=2
        )
        self.status_label.pack(fill=tk.X, padx=10)

        # Buttons frame
        buttons_frame = tk.Frame(self.root, bg='#1e1e1e')
        buttons_frame.pack(pady=5)

        # Open Excel button
        open_excel_btn = tk.Button(
            buttons_frame,
            text="📊 Open Excel",
            command=self.open_excel,
            font=('Consolas', 9, 'bold'),
            bg='#0e639c',
            fg='#ffffff',
            activebackground='#1177bb',
            activeforeground='#ffffff',
            cursor='hand2',
            padx=10,
            pady=5
        )
        open_excel_btn.pack(side=tk.LEFT, padx=5)

        # Open folder button
        open_folder_btn = tk.Button(
            buttons_frame,
            text="📁 Open Folder",
            command=self.open_folder,
            font=('Consolas', 9, 'bold'),
            bg='#0e639c',
            fg='#ffffff',
            activebackground='#1177bb',
            activeforeground='#ffffff',
            cursor='hand2',
            padx=10,
            pady=5
        )
        open_folder_btn.pack(side=tk.LEFT, padx=5)

        # Single-line entry field
        entry_frame = tk.Frame(self.root, bg='#1e1e1e')
        entry_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        entry_label = tk.Label(
            entry_frame,
            text=">",
            font=('Consolas', 14, 'bold'),
            bg='#1e1e1e',
            fg='#00ff00'
        )
        entry_label.pack(side=tk.LEFT, padx=(5, 5))

        self.entry_field = tk.Entry(
            entry_frame,
            font=('Consolas', 14),
            bg='#0c0c0c',
            fg='#00ff00',
            insertbackground='#00ff00',
            selectbackground='#264f78',
            relief=tk.FLAT,
            borderwidth=2
        )
        self.entry_field.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        self.entry_field.focus()

        # Bind Enter to submit
        self.entry_field.bind('<Return>', lambda e: self.submit_entry())

        # Set window colors
        self.root.configure(bg='#1e1e1e')

        # Handle window close - no confirmation
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def reset_timer(self):
        """Reset the timer to start a new 15-minute countdown"""
        self.timer_start_time = time.time()
        self.has_popped = False
        self.timer_running = True

    def start_timer_display(self):
        """Single timer thread that updates display and handles popup"""
        def timer_loop():
            while True:
                if not self.timer_running:
                    time.sleep(0.1)
                    continue

                elapsed = time.time() - self.timer_start_time
                remaining = (self.timer_minutes * 60) - elapsed

                if remaining <= 0:
                    if not self.has_popped:
                        self.pop_window()
                        self.has_popped = True
                    # Keep showing alert until user submits
                    self.timer_label.config(text="⏰ TIME TO LOG! ⏰", fg='#ff0000')
                else:
                    mins, secs = divmod(int(remaining), 60)
                    time_str = f"Next reminder in: {mins:02d}:{secs:02d}"
                    self.timer_label.config(text=time_str, fg='#ffff00')

                time.sleep(1)

        timer_thread = threading.Thread(target=timer_loop, daemon=True)
        timer_thread.start()

    def pop_window(self):
        """Bring window to foreground (pop up)"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        self.root.focus_force()
        self.entry_field.focus()

        # Flash the window to get attention
        self.root.state('normal')  # Ensure it's not minimized

    def parse_entry(self, line):
        """Parse a single entry line - position agnostic
        Accepts: "green $$ activity", "$$ g activity", "r activity $$", "g$$activity", etc.
        Color can be: green/g, red/r, white/w (with or without spaces)
        Returns: (color, dollars, activity) or None if invalid
        """
        line = line.strip()
        if not line:
            return None

        # Find color word or abbreviation (case insensitive, no word boundary)
        color_match = re.search(r'(green|red|white|g|r|w)', line, re.IGNORECASE)
        if not color_match:
            return None

        # Map abbreviations to full color names
        color_raw = color_match.group(1).lower()
        color_map = {'g': 'green', 'r': 'red', 'w': 'white'}
        color = color_map.get(color_raw, color_raw)

        # Find dollar signs
        dollars_match = re.search(r'\$+', line)
        if not dollars_match:
            return None
        dollars = dollars_match.group(0)

        # Remove color and dollars from line to get activity
        activity = line
        activity = re.sub(r'(green|red|white|g|r|w)', '', activity, flags=re.IGNORECASE)
        activity = re.sub(r'\$+', '', activity)
        activity = ' '.join(activity.split())  # Clean up extra whitespace

        if not activity:
            return None

        return (color, dollars, activity)

    def submit_entry(self):
        """Process and save entry to Excel"""
        text = self.entry_field.get().strip()

        if not text:
            return

        # Parse entry
        parsed = self.parse_entry(text)

        if not parsed:
            self.status_label.config(
                text='❌ Invalid format! Use: "green $$ activity" or "red $$$ activity" or "white $ activity"',
                fg='#ff0000'
            )
            return

        # Save to Excel
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active

            now = datetime.now()
            # Date format: 1jan2026
            date_str = now.strftime("%d%b%Y").lower()
            # Time format: 14:05
            time_str = now.strftime("%H:%M")

            color, dollars, activity = parsed

            row = [date_str, time_str, color.capitalize(), dollars, activity]
            ws.append(row)

            # Apply color to the entire row
            row_num = ws.max_row
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = self.color_fills[color]

            wb.save(self.excel_file)

            # Clear entry field
            self.entry_field.delete(0, tk.END)

            # Reset timer - single timer, just reset the start time
            self.reset_timer()

            # Show success with timestamp
            self.status_label.config(
                text=f'✓ Saved at {date_str} {time_str}: {color.upper()} {dollars} {activity}',
                fg='#00ff00'
            )

        except Exception as e:
            self.status_label.config(
                text=f'❌ Error: {str(e)}',
                fg='#ff0000'
            )

    def open_excel(self):
        """Open the Excel file with the default application"""
        if os.path.exists(self.excel_file):
            try:
                os.startfile(self.excel_file)
            except Exception as e:
                self.status_label.config(
                    text=f'❌ Failed to open Excel: {str(e)}',
                    fg='#ff0000'
                )
        else:
            self.status_label.config(
                text='❌ Excel file does not exist yet. Create an entry first!',
                fg='#ff0000'
            )

    def open_folder(self):
        """Open the folder containing the script and Excel file"""
        try:
            os.startfile(self.script_dir)
        except Exception as e:
            self.status_label.config(
                text=f'❌ Failed to open folder: {str(e)}',
                fg='#ff0000'
            )

    def on_closing(self):
        """Handle window close event - no confirmation"""
        self.timer_running = False
        self.root.destroy()
        sys.exit(0)

    def run(self):
        """Start the application"""
        self.root.mainloop()

if __name__ == "__main__":
    app = TimeTracker()
    app.run()
