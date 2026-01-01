#!/usr/bin/env python3
"""
Time & Energy Audit Tracker
Inspired by Dan Martell's time tracking strategy
Pops up every 15 minutes to log activities with color coding and dollar values

Copyright (c) 2026 Arty McLabin
"""

import tkinter as tk
from tkinter import messagebox
import threading
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import os
import re
import sys
import configparser

class TimeTracker:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("TimeAudit - 15:00")
        self.root.geometry("750x350")

        # Get script directory for Excel file and config
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_file = os.path.join(self.script_dir, "time_audit.xlsx")
        self.config_file = os.path.join(self.script_dir, "settings.ini")

        # Load settings
        self.timer_minutes = self.load_settings()

        self.setup_excel()

        # Timer setup - single timer, no duplication
        self.timer_start_time = None
        self.timer_running = False
        self.has_popped = False

        # Last entry tracking for prefill
        self.last_color = None
        self.last_dollars = None

        # Track last restore time to prevent immediate re-minimize
        self.last_restore_time = 0

        # Color mapping for Excel
        self.color_fills = {
            'green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'red': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
            'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }

        self.setup_ui()
        self.reset_timer()
        self.start_timer_display()

    def load_settings(self):
        """Load settings from .ini file, create with defaults if doesn't exist"""
        config = configparser.ConfigParser()

        if os.path.exists(self.config_file):
            config.read(self.config_file)
            timer_minutes = config.getint('Timer', 'interval_minutes', fallback=5)
        else:
            # Create default config
            timer_minutes = 5
            self.save_settings(timer_minutes)

        return timer_minutes

    def save_settings(self, timer_minutes):
        """Save settings to .ini file"""
        config = configparser.ConfigParser()
        config['Timer'] = {'interval_minutes': str(timer_minutes)}

        with open(self.config_file, 'w') as f:
            config.write(f)

    def setup_excel(self):
        """Initialize Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Time Audit"

            # Headers - Date, Day of Week, Time, Activity, Energy, Value, Interval
            headers = ['Date', 'Day', 'Time', 'Activity', 'Energy', 'Value', 'Interval']
            ws.append(headers)

            # Make headers bold
            for cell in ws[1]:
                cell.font = Font(bold=True)

            wb.save(self.excel_file)

    def setup_ui(self):
        """Create the UI"""
        # Modern dark background
        self.root.configure(bg='#1a1a2e')

        # Header with instructions
        header_frame = tk.Frame(self.root, bg='#16213e', relief=tk.FLAT)
        header_frame.pack(fill=tk.X, pady=(0, 10))

        instructions = tk.Label(
            header_frame,
            text='Instructions: github.com/ArtyMcLabin/TimeAuditTracker',
            font=('Segoe UI', 9),
            justify=tk.CENTER,
            bg='#16213e',
            fg='#a8a8a8',
            padx=15,
            pady=12,
            cursor='hand2'
        )
        instructions.pack()
        instructions.bind('<Button-1>', lambda e: os.startfile('https://github.com/ArtyMcLabin/TimeAuditTracker'))

        # Timer display with modern styling
        self.timer_label = tk.Label(
            self.root,
            text="Next reminder in: 15:00",
            font=('Segoe UI', 16, 'bold'),
            bg='#1a1a2e',
            fg='#4ecca3'
        )
        self.timer_label.pack(pady=(10, 5))

        # Timer interval selector with modern styling
        interval_frame = tk.Frame(self.root, bg='#1a1a2e')
        interval_frame.pack(pady=8)

        interval_label = tk.Label(
            interval_frame,
            text="Interval:",
            font=('Segoe UI', 10),
            bg='#1a1a2e',
            fg='#a8a8a8'
        )
        interval_label.pack(side=tk.LEFT, padx=(0, 8))

        self.interval_var = tk.StringVar(value=str(self.timer_minutes))
        interval_entry = tk.Entry(
            interval_frame,
            textvariable=self.interval_var,
            width=5,
            font=('Segoe UI', 10),
            bg='#16213e',
            fg='#e8e8e8',
            insertbackground='#4ecca3',
            selectbackground='#0f3460',
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=1,
            highlightbackground='#2a2a40',
            highlightcolor='#4ecca3'
        )
        interval_entry.pack(side=tk.LEFT, padx=5, ipady=4)
        interval_entry.bind('<Return>', self.on_interval_change)
        interval_entry.bind('<FocusOut>', self.on_interval_change)

        interval_suffix = tk.Label(
            interval_frame,
            text="min",
            font=('Segoe UI', 10),
            bg='#1a1a2e',
            fg='#a8a8a8'
        )
        interval_suffix.pack(side=tk.LEFT, padx=(8, 0))

        # Status label (shows save confirmation) with modern styling
        self.status_label = tk.Label(
            self.root,
            text="Ready to track...",
            font=('Segoe UI', 10),
            bg='#1a1a2e',
            fg='#a8a8a8',
            height=2
        )
        self.status_label.pack(fill=tk.X, padx=15, pady=(5, 0))

        # Buttons frame with modern styling
        buttons_frame = tk.Frame(self.root, bg='#1a1a2e')
        buttons_frame.pack(pady=10)

        # Modern button style
        btn_style = {
            'font': ('Segoe UI', 9),
            'bg': '#16213e',
            'fg': '#e8e8e8',
            'activebackground': '#0f3460',
            'activeforeground': '#4ecca3',
            'cursor': 'hand2',
            'relief': tk.FLAT,
            'borderwidth': 0,
            'padx': 15,
            'pady': 8
        }

        # Open Excel button
        open_excel_btn = tk.Button(
            buttons_frame,
            text="📊 Excel",
            command=self.open_excel,
            **btn_style
        )
        open_excel_btn.pack(side=tk.LEFT, padx=4)

        # Open folder button
        open_folder_btn = tk.Button(
            buttons_frame,
            text="📁 Folder",
            command=self.open_folder,
            **btn_style
        )
        open_folder_btn.pack(side=tk.LEFT, padx=4)

        # Reset timer button
        reset_timer_btn = tk.Button(
            buttons_frame,
            text="🔄 Reset",
            command=self.manual_reset_timer,
            **btn_style
        )
        reset_timer_btn.pack(side=tk.LEFT, padx=4)

        # Single-line entry field with modern styling
        entry_frame = tk.Frame(self.root, bg='#1a1a2e')
        entry_frame.pack(padx=20, pady=(10, 20), fill=tk.BOTH, expand=True)

        entry_label = tk.Label(
            entry_frame,
            text="❯",
            font=('Segoe UI', 18, 'bold'),
            bg='#1a1a2e',
            fg='#4ecca3'
        )
        entry_label.pack(side=tk.LEFT, padx=(0, 10))

        self.entry_field = tk.Entry(
            entry_frame,
            font=('Segoe UI', 14),
            bg='#16213e',
            fg='#e8e8e8',
            insertbackground='#4ecca3',
            selectbackground='#0f3460',
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=2,
            highlightbackground='#2a2a40',
            highlightcolor='#4ecca3'
        )
        self.entry_field.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, ipady=8)
        self.entry_field.focus()

        # Bind Enter to submit
        self.entry_field.bind('<Return>', lambda e: self.submit_entry())

        # Track focus events
        self.root.bind('<FocusIn>', self.on_focus_gained)
        self.root.bind('<FocusOut>', self.on_focus_lost)

        # Handle window close - no confirmation
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def reset_timer(self):
        """Reset the timer to start a new 15-minute countdown"""
        self.timer_start_time = time.time()
        self.has_popped = False
        self.timer_running = True

    def start_timer_display(self):
        """Single timer thread that updates display and handles popup"""
        def timer_loop():
            while True:
                if not self.timer_running:
                    time.sleep(0.1)
                    continue

                elapsed = time.time() - self.timer_start_time
                remaining = (self.timer_minutes * 60) - elapsed

                if remaining <= 0:
                    if not self.has_popped:
                        self.pop_window()
                        self.has_popped = True
                    # Keep showing alert until user submits
                    self.timer_label.config(text="⏰ TIME TO LOG! ⏰", fg='#e94560')
                    self.root.title("Clock Now!")
                else:
                    mins, secs = divmod(int(remaining), 60)
                    time_str = f"Next reminder in: {mins:02d}:{secs:02d}"
                    self.timer_label.config(text=time_str, fg='#4ecca3')
                    # Update window title with remaining time
                    self.root.title(f"TimeAudit - {mins:02d}:{secs:02d}")

                time.sleep(1)

        timer_thread = threading.Thread(target=timer_loop, daemon=True)
        timer_thread.start()

    def pop_window(self):
        """Bring window to foreground (pop up)"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        self.root.focus_force()
        self.entry_field.focus()

        # Flash the window to get attention
        self.root.state('normal')  # Ensure it's not minimized

    def parse_entry(self, line):
        """Parse a single entry line - position agnostic
        Accepts: "green $$ activity", "$$ g activity", "r activity $$", "activity g $$", etc.
        Color can be: green/g, red/r, white/w (anywhere in the string)
        Returns: (color, dollars, activity) or None if invalid
        """
        line = line.strip()
        if not line:
            return None

        # Find color with proper word boundaries
        # Full names need \b, single letters need to not be part of another word
        color_pattern = r'\b(green|red|white)\b|(?<![a-zA-Z])(g|r|w)(?![a-zA-Z])'
        color_match = re.search(color_pattern, line, re.IGNORECASE)
        if not color_match:
            return None

        # Get matched color from either group 1 or 2
        color_raw = (color_match.group(1) or color_match.group(2)).lower()
        color_map = {'g': 'green', 'r': 'red', 'w': 'white'}
        color = color_map.get(color_raw, color_raw)

        # Find dollar signs
        dollars_match = re.search(r'\$+', line)
        if not dollars_match:
            return None
        dollars = dollars_match.group(0)

        # Remove color and dollars from line to get activity
        # Remove the exact color match at its position (only once)
        activity = line[:color_match.start()] + line[color_match.end():]
        # Remove all dollar signs
        activity = re.sub(r'\$+', '', activity)
        activity = ' '.join(activity.split())  # Clean up extra whitespace

        if not activity:
            return None

        return (color, dollars, activity)

    def submit_entry(self):
        """Process and save entry to Excel"""
        text = self.entry_field.get().strip()

        if not text:
            return

        # Parse entry
        parsed = self.parse_entry(text)

        if not parsed:
            self.status_label.config(
                text='❌ Invalid format! Use: "g $$ activity" or "r $$$ activity" or "w $ activity"',
                fg='#e94560'
            )
            return

        # Save to Excel
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active

            now = datetime.now()
            # Date format: 1jan2026
            date_str = now.strftime("%d%b%Y").lower()
            # Day of week: Monday, Tuesday, etc.
            day_of_week = now.strftime("%A")
            # Time format: 14:05
            time_str = now.strftime("%H:%M")

            color, dollars, activity = parsed

            # Column order: Date, Day, Time, Activity, Energy, Value, Interval
            row = [date_str, day_of_week, time_str, activity, color.capitalize(), dollars, self.timer_minutes]
            ws.append(row)

            # Apply color to the entire row
            row_num = ws.max_row
            for col in range(1, 8):  # 7 columns total
                cell = ws.cell(row=row_num, column=col)
                cell.fill = self.color_fills[color]

            wb.save(self.excel_file)

            # Store last entry for prefill
            self.last_color = color
            self.last_dollars = dollars

            # Clear entry field and prefill with last color/dollars
            self.entry_field.delete(0, tk.END)

            # Prefill with abbreviated color and dollars
            color_abbrev = color[0]  # First letter: g, r, or w
            prefill = f"{color_abbrev} {dollars} "
            self.entry_field.insert(0, prefill)

            # Reset timer - single timer, just reset the start time
            self.reset_timer()

            # Show success with timestamp
            self.status_label.config(
                text=f'✓ Saved at {date_str} {time_str}: {color.upper()} {dollars} {activity}',
                fg='#4ecca3'
            )

        except Exception as e:
            self.status_label.config(
                text=f'❌ Error: {str(e)}',
                fg='#e94560'
            )

    def open_excel(self):
        """Open the Excel file with the default application"""
        if os.path.exists(self.excel_file):
            try:
                os.startfile(self.excel_file)
            except Exception as e:
                self.status_label.config(
                    text=f'❌ Failed to open Excel: {str(e)}',
                    fg='#e94560'
                )
        else:
            self.status_label.config(
                text='❌ Excel file does not exist yet. Create an entry first!',
                fg='#e94560'
            )

    def open_folder(self):
        """Open the folder containing the script and Excel file"""
        try:
            os.startfile(self.script_dir)
        except Exception as e:
            self.status_label.config(
                text=f'❌ Failed to open folder: {str(e)}',
                fg='#e94560'
            )

    def on_interval_change(self, event=None):
        """Handle timer interval change with validation"""
        try:
            new_interval = int(self.interval_var.get())

            # Validate range (1-999 minutes)
            if new_interval < 1 or new_interval > 999:
                raise ValueError("Interval must be between 1 and 999 minutes")

            self.timer_minutes = new_interval
            self.save_settings(new_interval)
            self.reset_timer()
            self.status_label.config(
                text=f'⏱️ Timer interval changed to {new_interval} minutes',
                fg='#f3a683'
            )
        except ValueError as e:
            # Invalid input, revert to previous value
            self.interval_var.set(str(self.timer_minutes))
            self.status_label.config(
                text=f'❌ Invalid interval: enter 1-999 minutes',
                fg='#e94560'
            )

    def on_focus_gained(self, event=None):
        """Track when window gains focus (restored from minimized)"""
        self.last_restore_time = time.time()

    def on_focus_lost(self, event=None):
        """Minimize window when clicking outside"""
        # Only minimize if it's not already minimized
        if self.root.state() != 'normal':
            return

        # Don't minimize if we just restored (within 0.5 seconds)
        time_since_restore = time.time() - self.last_restore_time
        if time_since_restore < 0.5:
            return

        # Minimize the window
        self.root.iconify()

    def manual_reset_timer(self):
        """Manually reset the timer countdown"""
        self.reset_timer()
        self.status_label.config(
            text='🔄 Timer manually reset',
            fg='#f3a683'
        )

    def on_closing(self):
        """Handle window close event - no confirmation"""
        self.timer_running = False
        self.root.destroy()
        sys.exit(0)

    def run(self):
        """Start the application"""
        self.root.mainloop()

if __name__ == "__main__":
    # Auto-close any existing instances
    import psutil
    current_pid = os.getpid()
    script_name = os.path.basename(__file__)

    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            # Check if it's a python process running this script
            if proc.info['cmdline'] and len(proc.info['cmdline']) > 1:
                if script_name in ' '.join(proc.info['cmdline']) and proc.info['pid'] != current_pid:
                    proc.kill()
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass

    app = TimeTracker()
    app.run()
